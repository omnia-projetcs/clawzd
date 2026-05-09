"""
Clawzd — Project Management module.
Provides Kanban / Table / Timeline views for project tracking.
Data stored as JSON files in data/projects/.
"""
import os, json, uuid, logging, asyncio
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from config import DATA_DIR

router = APIRouter()
logger = logging.getLogger("clawzd.project")

PROJECTS_DIR = os.path.join(DATA_DIR, "projects")
os.makedirs(PROJECTS_DIR, exist_ok=True)


# ── Helpers ──

def _proj_path(proj_id: str) -> str:
    return os.path.join(PROJECTS_DIR, f"{proj_id}.json")


def _load_proj(proj_id: str) -> Optional[dict]:
    p = _proj_path(proj_id)
    if os.path.exists(p):
        with open(p) as f:
            return json.load(f)
    return None


def _save_proj(proj: dict):
    proj["updated_at"] = datetime.now(timezone.utc).isoformat()
    with open(_proj_path(proj["id"]), "w") as f:
        json.dump(proj, f, indent=2, ensure_ascii=False)


def _list_projs() -> list:
    projs = []
    for fname in os.listdir(PROJECTS_DIR):
        if fname.endswith(".json"):
            try:
                with open(os.path.join(PROJECTS_DIR, fname)) as f:
                    data = json.load(f)
                    # Return summary only (no full tasks list for listing)
                    projs.append({
                        "id": data["id"],
                        "name": data.get("name", "Untitled"),
                        "description": data.get("description", ""),
                        "task_count": len(data.get("tasks", [])),
                        "created_at": data.get("created_at", ""),
                        "updated_at": data.get("updated_at", ""),
                    })
            except Exception:
                pass
    return sorted(projs, key=lambda w: w.get("updated_at", ""), reverse=True)


def _new_project(name: str = "New Project", description: str = "") -> dict:
    now = datetime.now(timezone.utc).isoformat()
    return {
        "id": str(uuid.uuid4())[:8],
        "name": name,
        "description": description,
        "created_at": now,
        "updated_at": now,
        "columns": ["Backlog", "To Do", "In Progress", "Review", "Done"],
        "members": [],
        "tasks": [],
        "github_repo": "",
        "github_project_id": "",
    }


# ── Project CRUD ──

@router.get("/projects")
async def list_projects():
    return {"projects": _list_projs()}


@router.post("/projects")
async def create_project(request: Request):
    data = await request.json()
    proj = _new_project(
        name=data.get("name", "New Project"),
        description=data.get("description", ""),
    )
    if "columns" in data:
        proj["columns"] = data["columns"]
    if "members" in data:
        proj["members"] = data["members"]
    _save_proj(proj)
    return {"status": "created", "project": proj}


@router.get("/projects/{proj_id}")
async def get_project(proj_id: str):
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    return {"project": proj}


@router.put("/projects/{proj_id}")
async def update_project(proj_id: str, request: Request):
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    data = await request.json()
    for k in ("name", "description", "columns", "members", "tasks",
              "github_repo", "github_project_id"):
        if k in data:
            proj[k] = data[k]
    _save_proj(proj)
    return {"status": "updated", "project": proj}


@router.delete("/projects/{proj_id}")
async def delete_project(proj_id: str):
    p = _proj_path(proj_id)
    if os.path.exists(p):
        os.remove(p)
    return {"status": "deleted"}


# ── Task CRUD ──

@router.post("/projects/{proj_id}/tasks")
async def add_task(proj_id: str, request: Request):
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    data = await request.json()
    now = datetime.now(timezone.utc).isoformat()
    task = {
        "id": "t" + str(uuid.uuid4())[:7],
        "title": data.get("title", "New Task"),
        "description": data.get("description", ""),
        "status": data.get("status", proj["columns"][0] if proj["columns"] else "Backlog"),
        "assignee": data.get("assignee", ""),
        "priority": data.get("priority", "medium"),
        "progress": data.get("progress", 0),
        "deadline": data.get("deadline", ""),
        "start_date": data.get("start_date", ""),
        "estimated_hours": data.get("estimated_hours", 0),
        "estimated_cost": data.get("estimated_cost", 0),
        "tags": data.get("tags", []),
        "created_at": now,
        "order": len(proj["tasks"]),
    }
    proj["tasks"].append(task)
    _save_proj(proj)
    return {"status": "created", "task": task}


@router.put("/projects/{proj_id}/tasks/{task_id}")
async def update_task(proj_id: str, task_id: str, request: Request):
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    data = await request.json()
    for t in proj["tasks"]:
        if t["id"] == task_id:
            for k in ("title", "description", "status", "assignee", "priority",
                       "progress", "deadline", "start_date", "estimated_hours",
                       "estimated_cost", "tags", "order"):
                if k in data:
                    t[k] = data[k]
            _save_proj(proj)
            return {"status": "updated", "task": t}
    raise HTTPException(404, "Task not found")


@router.delete("/projects/{proj_id}/tasks/{task_id}")
async def delete_task(proj_id: str, task_id: str):
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    proj["tasks"] = [t for t in proj["tasks"] if t["id"] != task_id]
    _save_proj(proj)
    return {"status": "deleted"}


@router.post("/projects/{proj_id}/reorder")
async def reorder_tasks(proj_id: str, request: Request):
    """Reorder tasks after drag & drop. Expects {task_id, new_status, new_order}."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    data = await request.json()
    task_id = data.get("task_id")
    new_status = data.get("new_status")
    new_order = data.get("new_order", 0)

    task = None
    for t in proj["tasks"]:
        if t["id"] == task_id:
            task = t
            break
    if not task:
        raise HTTPException(404, "Task not found")

    old_status = task["status"]
    task["status"] = new_status

    # Reorder within the target column
    col_tasks = [t for t in proj["tasks"] if t["status"] == new_status and t["id"] != task_id]
    col_tasks.sort(key=lambda x: x.get("order", 0))
    col_tasks.insert(min(new_order, len(col_tasks)), task)
    for i, ct in enumerate(col_tasks):
        ct["order"] = i

    _save_proj(proj)
    return {"status": "reordered"}


# ── Import TXT Todo ──

@router.post("/projects/{proj_id}/import/txt")
async def import_txt(proj_id: str, file: UploadFile = File(...)):
    """Import tasks from a .txt file (one task per line).
    Supports formats: plain lines, '- task', '* task', '- [ ] task', '- [x] task'.
    """
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    import re
    content = (await file.read()).decode("utf-8", errors="ignore")
    lines = content.splitlines()
    now = datetime.now(timezone.utc).isoformat()
    default_status = proj["columns"][0] if proj["columns"] else "Backlog"
    done_status = "Done" if "Done" in proj["columns"] else proj["columns"][-1] if proj["columns"] else "Done"
    imported = 0

    for line in lines:
        raw = line.strip()
        if not raw:
            continue

        # Detect done state from checkbox syntax: - [x] or * [x]
        is_done = False
        # Remove leading list markers: - [ ], - [x], * [ ], * [x], -, *, numbers
        cleaned = raw
        m = re.match(r'^[\-\*]\s*\[([ xX])\]\s*(.+)', cleaned)
        if m:
            is_done = m.group(1).lower() == 'x'
            cleaned = m.group(2).strip()
        else:
            cleaned = re.sub(r'^[\-\*\d\.]+\s*', '', cleaned).strip()

        if not cleaned:
            continue

        task = {
            "id": "t" + str(uuid.uuid4())[:7],
            "title": cleaned,
            "description": "",
            "status": done_status if is_done else default_status,
            "assignee": "",
            "priority": "medium",
            "progress": 100 if is_done else 0,
            "deadline": "",
            "start_date": "",
            "estimated_hours": 0,
            "estimated_cost": 0,
            "tags": [],
            "created_at": now,
            "order": len(proj["tasks"]),
        }
        proj["tasks"].append(task)
        imported += 1

    _save_proj(proj)
    return {"status": "imported", "imported": imported, "total_tasks": len(proj["tasks"])}


# ── AI Generation ──

@router.post("/ai-generate")
async def ai_generate_project(request: Request):
    """Use AI to generate a full project plan from a description."""
    data = await request.json()
    prompt = data.get("prompt", "")
    if not prompt:
        raise HTTPException(400, "Prompt is required")

    from app.llm_provider import get_llm_provider
    provider = get_llm_provider()

    system_prompt = """You are a project management AI assistant.
Given a project description, generate a complete project plan as JSON.

Return ONLY valid JSON matching this schema:
{
  "name": "Project Name",
  "description": "Brief description",
  "columns": ["Backlog", "To Do", "In Progress", "Review", "Done"],
  "members": ["Person1", "Person2"],
  "tasks": [
    {
      "id": "t1",
      "title": "Task title",
      "description": "Task description",
      "status": "To Do",
      "assignee": "Person1",
      "priority": "high",
      "progress": 0,
      "deadline": "2026-06-15",
      "start_date": "2026-05-15",
      "estimated_hours": 8,
      "estimated_cost": 400,
      "tags": ["design"],
      "order": 0
    }
  ]
}

Priority values: "critical", "high", "medium", "low"
Status values must be "To Do".
Generate 8-15 realistic tasks with varied priorities, and assignees.
Use realistic dates starting from today. Include logical task dependencies.
Assign tasks across team members. Estimate hours and costs realistically.
Return ONLY the JSON, no markdown fences, no extra text."""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Generate a project plan for: {prompt}"},
    ]

    try:
        response_text = ""
        async for chunk in provider.chat_stream(messages):
            response_text += chunk

        # Extract JSON
        start = response_text.find("{")
        end = response_text.rfind("}")
        if start != -1 and end != -1:
            json_str = response_text[start : end + 1]
            project_data = json.loads(json_str)

            # Ensure task IDs, order, and normalize dates/status
            today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            for i, task in enumerate(project_data.get("tasks", [])):
                if "id" not in task:
                    task["id"] = "t" + str(uuid.uuid4())[:7]
                task["order"] = i
                if "created_at" not in task:
                    task["created_at"] = datetime.now(timezone.utc).isoformat()
                # Force all tasks to today's date and "To Do" status
                task["start_date"] = today_str
                task["deadline"] = today_str
                task["status"] = "To Do"
                task["progress"] = 0

            return project_data
        else:
            raise ValueError("No JSON object found in AI response")
    except json.JSONDecodeError as e:
        logger.error(f"AI project generation JSON parse error: {e}")
        raise HTTPException(500, f"AI response was not valid JSON: {e}")
    except Exception as e:
        logger.error(f"AI project generation failed: {e}")
        raise HTTPException(500, f"AI generation failed: {e}")


# ── Export: Excel ──

@router.post("/projects/{proj_id}/export/excel")
async def export_excel(proj_id: str):
    """Export project tasks to an XLSX file."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = proj["name"][:31]  # Excel sheet name limit

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Title", "Status", "Assignee", "Priority", "Progress %",
               "Deadline", "Start Date", "Est. Hours", "Est. Cost ($)", "Tags"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Priority color mapping
    priority_fills = {
        "critical": PatternFill(start_color="FFCCCC", fill_type="solid"),
        "high": PatternFill(start_color="FFF3CD", fill_type="solid"),
        "medium": PatternFill(start_color="CCE5FF", fill_type="solid"),
        "low": PatternFill(start_color="E8E8E8", fill_type="solid"),
    }

    for row_idx, task in enumerate(proj["tasks"], 2):
        ws.cell(row=row_idx, column=1, value=task.get("title", ""))
        ws.cell(row=row_idx, column=2, value=task.get("status", ""))
        ws.cell(row=row_idx, column=3, value=task.get("assignee", ""))
        priority = task.get("priority", "medium")
        cell_p = ws.cell(row=row_idx, column=4, value=priority)
        if priority in priority_fills:
            cell_p.fill = priority_fills[priority]
        ws.cell(row=row_idx, column=5, value=task.get("progress", 0))
        ws.cell(row=row_idx, column=6, value=task.get("deadline", ""))
        ws.cell(row=row_idx, column=7, value=task.get("start_date", ""))
        ws.cell(row=row_idx, column=8, value=task.get("estimated_hours", 0))
        ws.cell(row=row_idx, column=9, value=task.get("estimated_cost", 0))
        ws.cell(row=row_idx, column=10, value=", ".join(task.get("tags", [])))
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    # Save to temp file
    export_dir = os.path.join(DATA_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    filename = f"project_{proj['name'][:20].replace(' ', '_')}_{proj['id']}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Export: Presentation ──

@router.post("/projects/{proj_id}/export/presentation")
async def export_presentation(proj_id: str):
    """Create a presentation in the Presentation Studio with project data."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    # Build a presentation data structure compatible with PresentationStudio
    pres_dir = os.path.join(DATA_DIR, "presentations")
    os.makedirs(pres_dir, exist_ok=True)

    pres_id = str(uuid.uuid4())[:8]
    now = datetime.now(timezone.utc).isoformat()

    # Summary stats
    total = len(proj["tasks"])
    done = len([t for t in proj["tasks"] if t.get("status") == "Done"])
    in_progress = len([t for t in proj["tasks"] if t.get("status") == "In Progress"])
    total_hours = sum(t.get("estimated_hours", 0) for t in proj["tasks"])
    total_cost = sum(t.get("estimated_cost", 0) for t in proj["tasks"])

    slides = []

    # Title slide
    slides.append({
        "id": "s1",
        "elements": [
            {"id": "e1", "type": "text", "x": 100, "y": 150, "w": 600, "h": 80,
             "content": proj["name"], "fontSize": 42, "fontWeight": "bold",
             "fill": "#ffffff", "textColor": "#ffffff"},
            {"id": "e2", "type": "text", "x": 100, "y": 250, "w": 600, "h": 40,
             "content": proj.get("description", ""), "fontSize": 18,
             "fill": "transparent", "textColor": "#a0a0a0"},
        ],
        "background": "#1a1a2e"
    })

    # Summary slide
    summary_text = (f"Total Tasks: {total}\n"
                    f"Completed: {done}\n"
                    f"In Progress: {in_progress}\n"
                    f"Estimated Hours: {total_hours}h\n"
                    f"Estimated Cost: ${total_cost}")
    slides.append({
        "id": "s2",
        "elements": [
            {"id": "e3", "type": "text", "x": 60, "y": 40, "w": 300, "h": 40,
             "content": "Project Summary", "fontSize": 28, "fontWeight": "bold",
             "fill": "transparent", "textColor": "#ffffff"},
            {"id": "e4", "type": "text", "x": 60, "y": 100, "w": 680, "h": 300,
             "content": summary_text, "fontSize": 18,
             "fill": "transparent", "textColor": "#e0e0e0"},
        ],
        "background": "#16213e"
    })

    # Task list slides (group by status)
    for col in proj.get("columns", []):
        col_tasks = [t for t in proj["tasks"] if t.get("status") == col]
        if not col_tasks:
            continue
        task_lines = "\n".join(
            f"• {t['title']} ({t.get('assignee', 'Unassigned')}) — {t.get('priority', 'medium')}"
            for t in col_tasks[:12]
        )
        sid = f"s{len(slides) + 1}"
        slides.append({
            "id": sid,
            "elements": [
                {"id": f"e{len(slides)*10+1}", "type": "text", "x": 60, "y": 40,
                 "w": 400, "h": 40, "content": col, "fontSize": 28,
                 "fontWeight": "bold", "fill": "transparent", "textColor": "#ffffff"},
                {"id": f"e{len(slides)*10+2}", "type": "text", "x": 60, "y": 100,
                 "w": 680, "h": 350, "content": task_lines, "fontSize": 14,
                 "fill": "transparent", "textColor": "#c0c0c0"},
            ],
            "background": "#0f3460"
        })

    pres = {
        "id": pres_id,
        "name": f"Project: {proj['name']}",
        "slides": slides,
        "created_at": now,
        "updated_at": now,
    }

    with open(os.path.join(pres_dir, f"{pres_id}.json"), "w") as f:
        json.dump(pres, f, indent=2, ensure_ascii=False)

    return {"status": "created", "presentation_id": pres_id, "name": pres["name"],
            "slide_count": len(slides)}


# ── GitHub Projects Sync ──

@router.post("/projects/{proj_id}/github/sync")
async def github_sync(proj_id: str, request: Request):
    """Sync project with GitHub Projects V2 (push local changes)."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    data = await request.json()
    token = data.get("token", os.environ.get("GITHUB_PAT", ""))
    repo = data.get("repo", proj.get("github_repo", ""))

    if not token:
        raise HTTPException(400, "GitHub PAT is required (set GITHUB_PAT in .env or provide in request)")
    if not repo:
        raise HTTPException(400, "GitHub repository (owner/repo) is required")

    import httpx

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    # Save repo to project
    proj["github_repo"] = repo
    _save_proj(proj)

    synced_issues = []
    errors = []

    async with httpx.AsyncClient(timeout=30) as client:
        for task in proj["tasks"]:
            # Map priority to labels
            labels = list(task.get("tags", []))
            priority = task.get("priority", "medium")
            labels.append(f"priority:{priority}")
            if task.get("status") == "Done":
                state = "closed"
            else:
                state = "open"

            body_parts = []
            if task.get("description"):
                body_parts.append(task["description"])
            body_parts.append(f"\n---\n**Assignee:** {task.get('assignee', 'Unassigned')}")
            body_parts.append(f"**Status:** {task.get('status', '')}")
            body_parts.append(f"**Priority:** {priority}")
            body_parts.append(f"**Progress:** {task.get('progress', 0)}%")
            if task.get("deadline"):
                body_parts.append(f"**Deadline:** {task['deadline']}")
            if task.get("estimated_hours"):
                body_parts.append(f"**Estimated Hours:** {task['estimated_hours']}h")
            if task.get("estimated_cost"):
                body_parts.append(f"**Estimated Cost:** ${task['estimated_cost']}")

            issue_data = {
                "title": task["title"],
                "body": "\n".join(body_parts),
                "labels": labels,
            }
            if task.get("assignee"):
                issue_data["assignees"] = [task["assignee"]]

            try:
                # Create issue via REST API
                resp = await client.post(
                    f"https://api.github.com/repos/{repo}/issues",
                    headers=headers,
                    json=issue_data,
                )
                if resp.status_code in (200, 201):
                    issue = resp.json()
                    synced_issues.append({
                        "task_id": task["id"],
                        "issue_number": issue["number"],
                        "url": issue["html_url"],
                    })
                    # Close if done
                    if state == "closed":
                        await client.patch(
                            f"https://api.github.com/repos/{repo}/issues/{issue['number']}",
                            headers=headers,
                            json={"state": "closed"},
                        )
                else:
                    errors.append({
                        "task_id": task["id"],
                        "error": f"HTTP {resp.status_code}: {resp.text[:200]}",
                    })
            except Exception as e:
                errors.append({"task_id": task["id"], "error": str(e)})

    return {
        "status": "synced",
        "synced": len(synced_issues),
        "errors": len(errors),
        "issues": synced_issues,
        "error_details": errors,
    }


@router.get("/projects/{proj_id}/github/import")
async def github_import(proj_id: str, request: Request):
    """Import issues from a GitHub repository into the project."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    token = request.query_params.get("token", os.environ.get("GITHUB_PAT", ""))
    repo = request.query_params.get("repo", proj.get("github_repo", ""))

    if not token or not repo:
        raise HTTPException(400, "GitHub PAT and repo are required")

    import httpx

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            f"https://api.github.com/repos/{repo}/issues",
            headers=headers,
            params={"state": "all", "per_page": 100},
        )
        if resp.status_code != 200:
            raise HTTPException(500, f"GitHub API error: {resp.status_code}")

        issues = resp.json()

    now = datetime.now(timezone.utc).isoformat()
    imported = 0

    for issue in issues:
        if issue.get("pull_request"):
            continue  # Skip PRs

        # Check if task already exists by title match
        existing = any(t["title"] == issue["title"] for t in proj["tasks"])
        if existing:
            continue

        # Map state to status
        if issue["state"] == "closed":
            status = "Done"
        else:
            # Check labels for status hints
            label_names = [l["name"].lower() for l in issue.get("labels", [])]
            if any("progress" in l or "wip" in l for l in label_names):
                status = "In Progress"
            elif any("review" in l for l in label_names):
                status = "Review"
            else:
                status = "To Do"

        # Extract priority from labels
        priority = "medium"
        for l in issue.get("labels", []):
            ln = l["name"].lower()
            if "critical" in ln:
                priority = "critical"
            elif "high" in ln:
                priority = "high"
            elif "low" in ln:
                priority = "low"

        task = {
            "id": "t" + str(uuid.uuid4())[:7],
            "title": issue["title"],
            "description": (issue.get("body") or "")[:500],
            "status": status,
            "assignee": issue["assignee"]["login"] if issue.get("assignee") else "",
            "priority": priority,
            "progress": 100 if status == "Done" else 0,
            "deadline": "",
            "start_date": issue.get("created_at", "")[:10],
            "estimated_hours": 0,
            "estimated_cost": 0,
            "tags": [l["name"] for l in issue.get("labels", []) if "priority" not in l["name"].lower()],
            "created_at": now,
            "order": len(proj["tasks"]),
        }
        proj["tasks"].append(task)
        imported += 1

    proj["github_repo"] = repo
    _save_proj(proj)

    return {"status": "imported", "imported": imported, "total_tasks": len(proj["tasks"])}


# ── Security Scan (Trivy) ──

@router.post("/projects/{proj_id}/scan/trivy")
async def trivy_scan(proj_id: str, request: Request):
    """Run Trivy security scan (fs + secret) on the project workspace."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    data = await request.json()
    scan_path = data.get("path", ".")  # default: workspace root
    scan_type = data.get("type", "fs")  # fs, secret, or both

    import subprocess, shutil

    # Check if Trivy is installed
    trivy_bin = shutil.which("trivy")
    if not trivy_bin:
        return {
            "status": "error",
            "message": "Trivy is not installed. Install it: curl -sfL https://raw.githubusercontent.com/aquasecurity/trivy/main/contrib/install.sh | sh -s -- -b /usr/local/bin",
            "results": []
        }

    # Resolve scan path
    from config import WORKSPACE_DIR
    abs_path = os.path.join(WORKSPACE_DIR, scan_path) if not os.path.isabs(scan_path) else scan_path
    if not os.path.exists(abs_path):
        raise HTTPException(400, f"Path not found: {scan_path}")

    results = []

    async def _run_trivy(mode: str):
        cmd = [trivy_bin, mode, "--format", "json", "--quiet", abs_path]
        try:
            proc = await asyncio.to_thread(
                subprocess.run, cmd, capture_output=True, text=True, timeout=120
            )
            if proc.stdout:
                return json.loads(proc.stdout)
        except subprocess.TimeoutExpired:
            return {"error": "Scan timed out after 120s"}
        except json.JSONDecodeError:
            return {"raw": proc.stdout[:2000] if proc.stdout else "", "stderr": proc.stderr[:500] if proc.stderr else ""}
        except Exception as e:
            return {"error": str(e)}
        return {}

    if scan_type in ("fs", "both"):
        fs_result = await _run_trivy("fs")
        results.append({"type": "filesystem", "data": fs_result})

    if scan_type in ("secret", "both"):
        secret_result = await _run_trivy("fs")  # trivy fs also scans secrets
        # Dedicated secret scan
        try:
            cmd = [trivy_bin, "fs", "--scanners", "secret", "--format", "json", "--quiet", abs_path]
            proc = await asyncio.to_thread(
                subprocess.run, cmd, capture_output=True, text=True, timeout=120
            )
            if proc.stdout:
                secret_data = json.loads(proc.stdout)
                results.append({"type": "secrets", "data": secret_data})
        except Exception as e:
            results.append({"type": "secrets", "error": str(e)})

    # Parse results into tasks
    vuln_tasks = []
    for r in results:
        data_block = r.get("data", {})
        if isinstance(data_block, dict):
            for res in data_block.get("Results", []):
                target = res.get("Target", "")
                for vuln in res.get("Vulnerabilities", []):
                    vuln_tasks.append({
                        "title": f"[{vuln.get('Severity','UNKNOWN')}] {vuln.get('VulnerabilityID','')}: {vuln.get('Title','')}",
                        "description": f"**Package:** {vuln.get('PkgName','')}\n**Installed:** {vuln.get('InstalledVersion','')}\n**Fixed:** {vuln.get('FixedVersion','N/A')}\n**File:** {target}\n\n{vuln.get('Description','')}",
                        "priority": _trivy_severity_to_priority(vuln.get("Severity", "")),
                        "tags": ["security", "trivy", vuln.get("Severity", "").lower()],
                    })
                for secret in res.get("Secrets", []):
                    vuln_tasks.append({
                        "title": f"[SECRET] {secret.get('RuleID','')}: {secret.get('Title','')}",
                        "description": f"**File:** {target}\n**Match:** `{secret.get('Match','')[:100]}...`\n**Category:** {secret.get('Category','')}",
                        "priority": "critical",
                        "tags": ["security", "secret", "trivy"],
                    })

    return {
        "status": "completed",
        "vulnerabilities": len(vuln_tasks),
        "tasks": vuln_tasks,
        "raw_results": results,
    }


def _trivy_severity_to_priority(severity: str) -> str:
    mapping = {"CRITICAL": "critical", "HIGH": "high", "MEDIUM": "medium", "LOW": "low"}
    return mapping.get(severity.upper(), "medium")


@router.post("/projects/{proj_id}/scan/trivy/apply")
async def trivy_apply_tasks(proj_id: str, request: Request):
    """Import Trivy scan results as project tasks."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    data = await request.json()
    tasks = data.get("tasks", [])
    now = datetime.now(timezone.utc).isoformat()
    default_col = "Backlog" if "Backlog" in proj.get("columns", []) else proj["columns"][0] if proj.get("columns") else "To Do"
    imported = 0

    for t in tasks:
        # Skip duplicates
        if any(existing["title"] == t["title"] for existing in proj["tasks"]):
            continue
        task = {
            "id": "t" + str(uuid.uuid4())[:7],
            "title": t["title"][:200],
            "description": t.get("description", ""),
            "status": default_col,
            "assignee": "",
            "priority": t.get("priority", "medium"),
            "progress": 0,
            "deadline": "",
            "start_date": "",
            "estimated_hours": 0,
            "estimated_cost": 0,
            "tags": t.get("tags", ["security"]),
            "created_at": now,
            "order": len(proj["tasks"]),
        }
        proj["tasks"].append(task)
        imported += 1

    _save_proj(proj)
    return {"status": "imported", "imported": imported, "total_tasks": len(proj["tasks"])}


# ── Internet Research for Project ──

@router.post("/projects/{proj_id}/research")
async def project_research(proj_id: str, request: Request):
    """Run internet research in parallel for a project — technologies, best practices, security advisories."""
    proj = _load_proj(proj_id)
    if not proj:
        raise HTTPException(404, "Project not found")

    data = await request.json()
    query = data.get("query", proj.get("name", ""))
    topics = data.get("topics", ["best practices", "security", "architecture", "tools"])

    from config import TAVILY_API_KEY

    async def _search(topic: str) -> dict:
        full_query = f"{query} {topic}"
        all_results = []

        async def _tavily():
            if not TAVILY_API_KEY:
                return []
            try:
                from tavily import AsyncTavilyClient
                client = AsyncTavilyClient(api_key=TAVILY_API_KEY)
                response = await client.search(full_query, max_results=5)
                return [
                    {"title": r.get("title", ""), "snippet": r.get("content", ""),
                     "url": r.get("url", ""), "source": "tavily"}
                    for r in response.get("results", [])
                ]
            except Exception as e:
                logger.warning("Tavily search failed for '%s': %s", topic, e)
                return []

        async def _ddg():
            try:
                from ddgs import DDGS
                results = await asyncio.to_thread(
                    lambda: list(DDGS().text(full_query, max_results=5))
                )
                return [
                    {"title": r.get("title", ""), "snippet": r.get("body", ""),
                     "url": r.get("href", ""), "source": "duckduckgo"}
                    for r in results
                ]
            except Exception as e:
                logger.warning("DDG search failed for '%s': %s", topic, e)
                return []

        # Run both in parallel
        tavily_r, ddg_r = await asyncio.gather(_tavily(), _ddg())

        # Merge & deduplicate
        seen = set()
        for r in tavily_r + ddg_r:
            url = r.get("url", "")
            if url and url not in seen:
                seen.add(url)
                all_results.append(r)

        return {"topic": topic, "results": all_results}

    # Run all searches in parallel
    search_tasks = [_search(t) for t in topics[:6]]
    results = await asyncio.gather(*search_tasks)

    # Synthesize results with LLM
    all_results = []
    for r in results:
        all_results.extend(r.get("results", []))

    summary = ""
    suggested_tasks = []

    if all_results:
        try:
            from app.llm_provider import get_llm_provider
            provider = get_llm_provider()

            results_text = "\n".join(
                f"- [{r['title']}]({r['url']}): {r['snippet'][:150]}"
                for r in all_results[:20]
            )

            messages = [
                {"role": "system", "content": (
                    "You are a project analyst. Analyze these research results and provide:\n"
                    "1. A brief summary of key findings\n"
                    "2. Suggested action items as tasks\n\n"
                    "Return JSON: {\"summary\": \"...\", \"tasks\": [{\"title\": \"...\", \"description\": \"...\", \"priority\": \"high|medium|low\", \"tags\": [\"...\"]}]}\n"
                    "Return ONLY valid JSON."
                )},
                {"role": "user", "content": f"Project: {query}\n\nResearch results:\n{results_text}"}
            ]

            response_text = ""
            async for chunk in provider.chat_stream(messages):
                response_text += chunk

            start = response_text.find("{")
            end = response_text.rfind("}")
            if start != -1 and end != -1:
                parsed = json.loads(response_text[start:end+1])
                summary = parsed.get("summary", "")
                suggested_tasks = parsed.get("tasks", [])
        except Exception as e:
            logger.warning("LLM synthesis failed: %s", e)
            summary = f"Found {len(all_results)} results across {len(topics)} topics."

    return {
        "status": "completed",
        "search_results": results,
        "summary": summary,
        "suggested_tasks": suggested_tasks,
        "total_results": len(all_results),
    }
